# window create
from tkinter import *
from tkinter.ttk import *
from tkinter import scrolledtext, messagebox
from tkinter import ttk
import os
from openpyxl import Workbook, load_workbook

# Main window
window = Tk()
window.title("The Form")
window.geometry("800x600")

# Create Notebook
my_notebook = ttk.Notebook(window)
my_notebook.pack(expand=1, fill=BOTH)


# function to register and save to excel
def reg():
    stu = entry1.get().strip()
    reg = entry2.get().strip()
    mark1 = entry3.get().strip()
    mark2 = entry4.get().strip()
    mark3 = entry5.get().strip()
    mark4 = entry6.get().strip()
    mark5 = entry7.get().strip()

    if not all([stu, reg, mark1, mark2, mark3, mark4, mark5]):
        messagebox.showwarning("Input Error", "All feilds are required")
        return
    file_name = "student_data5.xlsx"
    if not os.path.exists(file_name):
        # create workbook and sheet with reader
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Data"
        ws.append(["Student Name", "Register No",
                   "Tamil", "English", "Maths", "Science", "Social Science",
                   "Total", "Average", "Result"
                   ])
    else:
        # load exing workbook
        wb = load_workbook(file_name)
        ws = wb.active

    # Convert marks to integers before performing calculations
    mark1 = int(mark1) if mark1.isdigit() else 0
    mark2 = int(mark2) if mark2.isdigit() else 0
    mark3 = int(mark3) if mark3.isdigit() else 0
    mark4 = int(mark4) if mark4.isdigit() else 0
    mark5 = int(mark5) if mark5.isdigit() else 0

    tot = mark1 + mark2 + mark3 + mark4 + mark5
    avg = tot / 5

    # Check pass or fail
    if mark1 >= 35 and mark2 >= 35 and mark3 >= 35 and mark4 >= 35 and mark5 >= 35:
        result = "Pass"
    else:
        result = "Fail"

    # Save the data to Excel
    ws.append([stu, reg, mark1, mark2, mark3, mark4, mark5, tot, avg, result])
    wb.save(file_name)

    # Print the results
    print("\n!!************* Mark Sheet ************!!")
    print("Student Name:", stu)
    print("Register No.:", reg)
    print("Tamil:", mark1)
    print("English:", mark2)
    print("Maths:", mark3)
    print("Science:", mark4)
    print("Social Science:", mark5)
    print("Total Marks:", tot)
    print("Average Marks:", avg)
    print("Result:", result)
    print("\n")

    # Clear the entry box
    entry1.delete(0, END)
    entry2.delete(0, END)
    entry3.delete(0, END)
    entry4.delete(0, END)
    entry5.delete(0, END)
    entry6.delete(0, END)
    entry7.delete(0, END)

    messagebox.showinfo("Succes", "Student data saved to excel!")


# Tab frame
tab1 = ttk.Frame(my_notebook)
my_notebook.add(tab1, text="Student Form")

lab0 = Label(tab1, text="STUDENT Information:", foreground="Green", font=("Capri", 12, "bold"))
lab0.grid(row=1, column=8, pady=10)

# Label 1
lab = Label(tab1, text="STUDENT NAME:", foreground="black", font=("Capri", 12, "bold"))
lab.grid(row=2, column=4, pady=10)

entry1 = Entry(tab1, width=30)
entry1.grid(row=2, column=6, padx=20, pady=10)

# Label 2
lab2 = Label(tab1, text="REG NO:", foreground="black", font=("Capri", 12, "bold"))
lab2.grid(row=4, column=4, pady=10)

entry2 = Entry(tab1, width=30)
entry2.grid(row=4, column=6, padx=10, pady=10)

# Label 3
lab3 = Label(tab1, text="TAMIL:", foreground="black", font=("Capri", 12, "bold"))
lab3.grid(row=6, column=4, pady=10)

entry3 = Entry(tab1, width=30)
entry3.grid(row=6, column=6, padx=10, pady=10)

# Label 4
lab4 = Label(tab1, text="ENGLISH:", foreground="black", font=("Capri", 12, "bold"))
lab4.grid(row=8, column=4, pady=10)

entry4 = Entry(tab1, width=30)
entry4.grid(row=8, column=6, padx=10, pady=10)

# Label 5
lab5 = Label(tab1, text="MATHS:", foreground="black", font=("Capri", 12, "bold"))
lab5.grid(row=10, column=4, pady=10)

entry5 = Entry(tab1, width=30)
entry5.grid(row=10, column=6, padx=10, pady=10)

# Label 6
lab6 = Label(tab1, text="SCIENCE:", foreground="black", font=("Capri", 12, "bold"))
lab6.grid(row=12, column=4, pady=10)

entry6 = Entry(tab1, width=30)
entry6.grid(row=12, column=6, padx=10, pady=10)

# Label 7
lab7 = Label(tab1, text="SOCIAL SCIENCE:", foreground="black", font=("Capri", 12, "bold"))
lab7.grid(row=14, column=4, pady=10)

entry7 = Entry(tab1, width=30)
entry7.grid(row=14, column=6, padx=10, pady=10)

# Submit button
submit_btn = Button(tab1, text="Submit", command=reg)
submit_btn.grid(row=16, column=6, padx=20, pady=20)

# create workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Student Data"

# add column headers
ws.append([
    "Student Name", "Register No",
    "Tamil", "English", "Maths", "Science", "Social Science",
    "Total", "Average", "Result"
])

# Save file only after closing the program
wb.save("student_data5.xlsx")
print("Excel file 'student_data5.xlsx' created successfully!")
